import openpyxl
import os

def create_excel_with_headers(file_path, sheet_name, headers, data_row=None):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    if data_row:
        sheet.append(data_row)
    workbook.save(file_path)

def create_account_base_excel(file_path):
    workbook = openpyxl.Workbook()

    # Level sheet
    level_sheet = workbook.active
    level_sheet.title = "层级信息"
    level_sheet.append(["NSC_id", "第二期层级"])
    level_sheet.append(["NSC001", "A"])

    # Store sheet
    store_sheet = workbook.create_sheet(title="门店信息")
    store_sheet.append(["NSC Code", "抖音id"])
    store_sheet.append(["NSC001", "douyin_store_1"])

    workbook.save(file_path)

if __name__ == "__main__":
    data_dir = "data"
    os.makedirs(data_dir, exist_ok=True)

    # Define files and their headers based on source_mappings.py
    files_config = {
        "video_data.xlsx": {
            "sheet_name": "Sheet1",
            "headers": ["主机厂经销商id", "日期", "锚点曝光次数", "锚点点击次数", "新发布视频数", "短视频表单提交商机量"],
            "data_row": ["NSC001", "2025-01-01", 100, 10, 5, 2]
        },
        "live_data.xlsx": {
            "sheet_name": "Sheet1",
            "headers": ["主机厂经销商id列表", "开播日期", "超25分钟直播时长(分)", "直播有效时长（小时）", "超25min直播总场次", "曝光人数", "场观", "小风车点击次数（不含小雪花）"],
            "data_row": ["NSC001", "2025-01-01", 60, 1.0, 1, 500, 100, 50]
        },
        "msg_data.xlsx": {
            "sheet_name": "2025-01-01", # Sheet name with date
            "headers": ["主机厂经销商ID", "日期", "进入私信客户数", "主动咨询客户数", "私信留资客户数"],
            "data_row": ["NSC001", "2025-01-01", 5, 3, 1]
        },
        "account_bi.xlsx": {
            "sheet_name": "Sheet1",
            "headers": ["主机厂经销商id列表", "日期", "直播间表单提交商机量", "短-播放量"],
            "data_row": ["NSC001", "2025-01-01", 10, 200]
        },
        "leads.xlsx": {
            "sheet_name": "Sheet1",
            "headers": ["主机厂经销商id列表", "留资日期", "直播间表单提交商机量(去重)"],
            "data_row": ["NSC001", "2025-01-01", 3]
        },
        "spending.xlsx": {
            "sheet_name": "Sheet1",
            "headers": ["NSC CODE", "Date", "Spending(Net)"],
            "data_row": ["NSC001", "2025-01-01", 1000.0]
        },
    }

    for filename, config in files_config.items():
        file_path = os.path.join(data_dir, filename)
        create_excel_with_headers(file_path, config["sheet_name"], config["headers"], config["data_row"])
        print(f"Created {file_path} with headers and data.")

    # Special handling for account_base.xlsx
    account_base_path = os.path.join(data_dir, "account_base.xlsx")
    create_account_base_excel(account_base_path)
    print(f"Created {account_base_path} with multiple sheets, headers and data.")